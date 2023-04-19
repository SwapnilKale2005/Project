from django.shortcuts import render
from django.http import HttpResponse
import mysql.connector 
import openpyxl
import io
from home.models import Course
from home.models import Marks
from django.db import connection, connections

def home(request):
    return render(request,'home.html')

def excel(request):
    if request.method == 'POST':
        
        course_code = request.POST['course_code']
        branch = request.POST.get('branch')
        year = request.POST.get('year')
       
        
   
        # Select the active worksheet
        #********************storing marks***************************** 
        # Define the columns to be fetched
        columns = ['roll_no', 'Question1', 'Question2', 'Question3', 'Question4', 'Question5', 'Question6', 'Question7', 'Question8']

        # Construct the SQL query
        query = "SELECT {} FROM m{} WHERE branch = %s AND year = %s ORDER BY roll_no, year".format(", ".join(columns), course_code)

        # Execute the query and fetch the results
        with connection.cursor() as cursor:
            cursor.execute(query, [branch, year])
            marks = cursor.fetchall()

        # Write the results to Excel cells
        file_path='static\PT2excel.xlsx'
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for row_num, row_data in enumerate(marks):
            for col_num, cell_value in enumerate(row_data):
                    ws.cell(row=row_num+12, column=col_num+2).value = cell_value
                    
        # Save the workbook
        wb.save('static/PT2excel.xlsx')
      
        #********************************END*****************************
        #**************************upbadting cource outcome***************

        # Define the columns to be fetched
        columns = ['question1', 'question2', 'question3', 'question4', 'question5', 'question6', 'question7', 'question8']

        # Construct the SQL query
        query2 = "SELECT {} FROM c{} WHERE branch = %s AND year = %s ".format(", ".join(columns), course_code)

        # Execute the query and fetch the results
        with connection.cursor() as cursor:
            cursor.execute(query2, [branch, year])
            courseOutcome = cursor.fetchall()

        # Write the results to Excel cells
        for rownum, rowdata in enumerate(courseOutcome):
            for colnum, cellvalue in enumerate(rowdata):
                ws.cell(row=rownum+10, column=colnum+3).value = cellvalue
               

        # Save the workbook
        wb.save('static/PT2excel.xlsx')
       
        #********************************END*****************************


        return render(request,'download.html')
    return render(request,'excel.html')
    

def remove(request):
    if request.method == 'POST':
        wb = openpyxl.load_workbook('static\PT2excel.xlsx')
   
        # Select the active worksheet
        ws = wb.active
        for row in ws.iter_rows(min_row=12, min_col=2, max_row=80, max_col=10):
                for cell in row:
                        cell.value = None
        wb.save('static\PT2excel.xlsx')
        return render(request,'home.html')
    
                        

def about(request):
    return render(request,'about.html')

def setCourceOutcome(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')
        #code for storng course outcome
        Q1= request.POST.get('Q1')
        Q2= request.POST.get('Q2')
        Q3= request.POST.get('Q3')
        Q4= request.POST.get('Q4')
        Q5= request.POST.get('Q5')
        Q6= request.POST.get('Q6')
        Q7= request.POST.get('Q7')
        Q8= request.POST.get('Q8')

        # Example data to be inserted into the database
        data = {
            'branch':branch,
            'year':year,
            'pt':pt,
            'question1': Q1,
            'question2': Q2,
            'question3': Q3,
            'question4': Q4,
            'question5': Q5,
            'question6': Q6,
            'question7': Q7,
            'question8': Q8,
    
            }
        #******************Insert data into the  course tabel*************
        def insert_data_into_table(table_name, data):
            with connection.cursor() as cursor:
                        # Construct the INSERT query
                        columns = ', '.join(data.keys())
                        placeholders = ', '.join(['%s'] * len(data))
                        query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"

                        # Execute the query with the given data
                        cursor.execute(query, list(data.values()))

            # Commit the changes to the database
            connection.commit()

        # Call the function to insert the data into the table
        insert_data_into_table('c'+course_code, data)
        
                      
        return render(request,'courseSetSuccessfully.html')

    return render(request,'setCourse.html')
 
def setPaper(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')

        #*********************Creating course tabel*************************
        # Check if the table exists
        existing_tables = connection.introspection.table_names()
        if 'c'+course_code in existing_tables:
            with connections['default'].cursor() as cursor:
                cursor.execute(f"DROP TABLE c{course_code}")
        
        # Create a dynamic model class based on the DynamicTable model
        class NewCourseOutcome(Course):
            class Meta:
                db_table = 'c'+course_code
        
        # Save the dynamic model class in a module-level variable
        globals()[NewCourseOutcome.__name__] = NewCourseOutcome

        # Create the new table
        with connection.schema_editor() as schema_editor:
            schema_editor.create_model(NewCourseOutcome)
        #***********************END*****************************

        #*********************Creating marks tabel*************************
        existing_tables = connection.introspection.table_names()
        if 'm'+course_code in existing_tables:
            with connections['default'].cursor() as cursor:
                cursor.execute(f"DROP TABLE m{course_code}")
              

        # Create a dynamic model class based on the DynamicTable model
        class NewMarks(Marks):
            class Meta:
                db_table = 'm'+course_code
        
        # Save the dynamic model class in a module-level variable
        globals()[NewCourseOutcome.__name__] = NewMarks
        
        # Create the new table
        with connection.schema_editor() as schema_editor:
            schema_editor.create_model(NewMarks)
        #***********************END*****************************

        return render(request,'setPaperSuccessfully.html')

    return render(request,'setPaper.html')

def displayPaper(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')
        # Define the columns to be fetched
        columns = ['question1', 'question2', 'question3', 'question4', 'question5', 'question6', 'question7', 'question8']

        # Construct the SQL query
        query2 = "SELECT {} FROM c{} WHERE branch = %s AND year = %s ".format(", ".join(columns), course_code)

        # Execute the query and fetch the results
        with connection.cursor() as cursor:
            cursor.execute(query2, [branch, year])
            courseOutcome = cursor.fetchall()
 
            resultArray = [courseOutcome[0][:3],courseOutcome[0][3:8]]
            letterArray =['A','B','C','D','E']
        
        return render(request,'displayPaperSuccessfully.html',{'resultArray':resultArray,'length_range': length_range})
    return render(request,'displayPaper.html')

def length_range(start, end):
    return range(start, end + 1)

def capture(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        roll_no= request.POST.get('roll_no')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')
        print(course_code)
        print(roll_no)
        print(branch)
        print(year)
       
        query = f"INSERT INTO m{course_code} (roll_no, branch, year, pt, question1, question2, question3, question4, question5, question6, question7, question8) VALUES ('{roll_no}', '{branch}', '{year}' , '{pt}' ,'0' ,'0' ,'0' ,'0' ,'0' ,'0' ,'0' ,'0') "

        # Execute the SQL query
        with connection.cursor() as cursor:
            cursor.execute(query)

        return render(request,'capture.html')
    return render(request,'studentLogin.html')

def storeMarks(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        roll_no= request.POST.get('roll_no')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')
        #code for storng course outcome
        Q1= request.POST.get('Q1')
        Q2= request.POST.get('Q2')
        Q3= request.POST.get('Q3')
        Q4= request.POST.get('Q4')
        Q5= request.POST.get('Q5')
        Q6= request.POST.get('Q6')
        Q7= request.POST.get('Q7')
        Q8= request.POST.get('Q8')
        print(course_code)
        print(roll_no)
        print(branch)
        print(year)
        print(pt)
        print(Q1)
        print(Q2)
        print(Q3)
        print(Q4)
        print(Q5)
        print(Q6)
        print(Q7)
        print(Q8)
        

         # Construct the SQL query to update the table
        query = f"UPDATE m{course_code} SET pt='{pt}', question1='{Q1}', question2='{Q2}', question3='{Q3}', question4='{Q4}', question5='{Q5}', question6='{Q6}', question7='{Q7}', question8='{Q8}' WHERE roll_no='{roll_no}';"

        # Execute the SQL query
        with connection.cursor() as cursor:
            cursor.execute(query)

       
                      
        return render(request,'storeMarksSuccessfully.html')

    return render(request,'capture.html')

    
        
